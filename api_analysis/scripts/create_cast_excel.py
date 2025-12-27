#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ReelShort日本オリジナル作品のキャスト情報をExcelにまとめるスクリプト
ブラウザ検索で収集した情報を統合
"""

import csv
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 収集したキャスト情報データ
CAST_DATA = {
    "冷酷御曹司の愛妻計画": {
        "status": "✅確認済",
        "production": "nowhere film / NISSIN / ベルテ",
        "source": "PR TIMES (2025/5/12)",
        "cast": [
            {"name": "長田成哉", "role": "高嶺総一郎", "agency": "不明", "social": "映画「変な家」出演"},
            {"name": "結城陽葵", "role": "篠宮凛/吉野有紗", "agency": "不明", "social": "『続・続・最後から二番目の恋』出演"},
            {"name": "福田博之", "role": "篠宮慎", "agency": "不明", "social": ""},
            {"name": "関塚まいこ", "role": "篠宮純子", "agency": "不明", "social": ""},
            {"name": "新羅美玲", "role": "篠宮萌香", "agency": "不明", "social": ""},
            {"name": "宮山典子", "role": "篠宮小百合", "agency": "不明", "social": ""},
            {"name": "加藤エン", "role": "平雄介", "agency": "不明", "social": ""},
            {"name": "清なをみ", "role": "吉野晴子", "agency": "不明", "social": ""},
            {"name": "大山かりん", "role": "間宮優香", "agency": "不明", "social": ""},
            {"name": "小川ガオ", "role": "水野肇", "agency": "不明", "social": ""},
        ]
    },
    "財閥令嬢様の二重生活": {
        "status": "✅確認済",
        "production": "nowhere film",
        "source": "nowhere-film.jp",
        "cast": [
            {"name": "音野暁", "role": "主演", "agency": "不明", "social": ""},
            {"name": "江畑浩規", "role": "出演", "agency": "不明", "social": ""},
            {"name": "梁錦川", "role": "出演", "agency": "不明", "social": ""},
        ]
    },
    "一目惚れ！今すぐ結婚してくれますか？": {
        "status": "✅確認済",
        "production": "不明",
        "source": "miyanaoko.com",
        "cast": [
            {"name": "みやなおこ", "role": "出演", "agency": "officeNAO", "social": "Instagram: miyanaoko.com, Web: https://miyanaoko.com/"},
        ]
    },
    "帰ってきたお嬢様": {
        "status": "✅確認済",
        "production": "不明",
        "source": "Instagram (eyes inc)",
        "cast": [
            {"name": "西野莉世", "role": "神代いずみ（主演）", "agency": "eyes inc", "social": "Instagram: @nishino_rise, 2002年4月9日生まれ、東京都出身"},
        ]
    },
    "禁断の誘い～彼のものに～": {
        "status": "✅確認済",
        "production": "不明",
        "source": "bianchetto.co.jp",
        "cast": [
            {"name": "新海航輝", "role": "石山亮平", "agency": "Bianchetto", "social": ""},
            {"name": "秋葉七海", "role": "社員役", "agency": "Bianchetto", "social": ""},
        ]
    },
    "せめて最後に愛のキスを": {
        "status": "✅確認済",
        "production": "不明",
        "source": "919quick.com / Instagram",
        "cast": [
            {"name": "紀乃はる", "role": "出演", "agency": "株式会社Quick", "social": ""},
            {"name": "宮野翼", "role": "克也の秘書", "agency": "不明", "social": "Instagram: miyano_tsubasa"},
        ]
    },
    "奪われたプリマバレリーナ": {
        "status": "✅確認済",
        "production": "不明",
        "source": "Instagram (@satou617_y)",
        "cast": [
            {"name": "satou617_y", "role": "小川里奈", "agency": "不明", "social": "Instagram: @satou617_y"},
        ]
    },
    "もう一度、君に恋をする": {
        "status": "❌未確認",
        "production": "不明",
        "source": "",
        "cast": []
    },
    "消防士の元夫、悔恨の炎に焼かれて": {
        "status": "❌未確認",
        "production": "不明",
        "source": "",
        "cast": []
    },
    "無職の夫は大富豪だった件!": {
        "status": "✅確認済",
        "production": "不明",
        "source": "Instagram (@kato_sho_)",
        "cast": [
            {"name": "加藤将", "role": "織田克宏（主演）", "agency": "不明", "social": "Instagram: @kato_sho_"},
        ]
    },
    "ダイヤモンドの再会": {
        "status": "❌未確認",
        "production": "不明",
        "source": "",
        "cast": []
    },
    "銃と金とクリスマス": {
        "status": "❌未確認",
        "production": "不明",
        "source": "",
        "cast": []
    },
    "ケーキ職人のワタシが大富豪と偽装結婚した話": {
        "status": "❌未確認",
        "production": "不明",
        "source": "",
        "cast": []
    },
    "Yesから始まるラブストーリー": {
        "status": "❌未確認",
        "production": "不明",
        "source": "",
        "cast": []
    },
    "負け犬の私がバージンを捨てる！": {
        "status": "❌未確認",
        "production": "不明",
        "source": "",
        "cast": []
    },
    "嫌いなアイツの専属メイド!?": {
        "status": "❌未確認",
        "production": "不明",
        "source": "",
        "cast": []
    },
    "令嬢決戦！私こそが学園のクイーン": {
        "status": "❌未確認",
        "production": "不明",
        "source": "",
        "cast": []
    },
}

def create_excel():
    """Excelファイルを作成"""
    wb = Workbook()
    
    # スタイル定義
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    subheader_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    found_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    not_found_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ===== シート1: サマリー =====
    ws1 = wb.active
    ws1.title = "サマリー"
    
    ws1.merge_cells('A1:G1')
    ws1['A1'] = "ReelShort 日本オリジナル作品 キャスト情報調査レポート"
    ws1['A1'].font = Font(bold=True, size=16)
    ws1['A1'].alignment = Alignment(horizontal="center")
    
    ws1['A3'] = f"作成日時: {datetime.now().strftime('%Y年%m月%d日 %H:%M')}"
    ws1['A4'] = "調査方法: DuckDuckGo検索、俳優事務所サイト、PR TIMES等"
    
    # 統計
    total = len(CAST_DATA)
    confirmed = sum(1 for v in CAST_DATA.values() if "✅" in v["status"])
    total_actors = sum(len(v["cast"]) for v in CAST_DATA.values())
    
    ws1['A6'] = "調査統計"
    ws1['A6'].font = Font(bold=True, size=12)
    ws1['A7'] = f"総作品数: {total}"
    ws1['A8'] = f"確認済み作品数: {confirmed}"
    ws1['A9'] = f"未確認作品数: {total - confirmed}"
    ws1['A10'] = f"確認済み俳優数: {total_actors}"
    ws1['A11'] = f"進捗率: {confirmed/total*100:.1f}%"
    
    # ===== シート2: 作品別キャスト =====
    ws2 = wb.create_sheet("作品別キャスト")
    
    headers = ["作品名", "調査状況", "制作会社", "俳優名", "役名", "事務所", "SNS/備考", "情報ソース"]
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    row = 2
    for title, data in CAST_DATA.items():
        if data["cast"]:
            for i, actor in enumerate(data["cast"]):
                ws2.cell(row=row, column=1, value=title if i == 0 else "").border = thin_border
                ws2.cell(row=row, column=2, value=data["status"] if i == 0 else "").border = thin_border
                ws2.cell(row=row, column=3, value=data["production"] if i == 0 else "").border = thin_border
                ws2.cell(row=row, column=4, value=actor["name"]).border = thin_border
                ws2.cell(row=row, column=5, value=actor["role"]).border = thin_border
                ws2.cell(row=row, column=6, value=actor["agency"]).border = thin_border
                ws2.cell(row=row, column=7, value=actor["social"]).border = thin_border
                ws2.cell(row=row, column=8, value=data["source"] if i == 0 else "").border = thin_border
                
                fill = found_fill if "✅" in data["status"] else not_found_fill
                for col in range(1, 9):
                    ws2.cell(row=row, column=col).fill = fill
                row += 1
        else:
            ws2.cell(row=row, column=1, value=title).border = thin_border
            ws2.cell(row=row, column=2, value=data["status"]).border = thin_border
            ws2.cell(row=row, column=3, value=data["production"]).border = thin_border
            for col in range(4, 9):
                ws2.cell(row=row, column=col, value="-").border = thin_border
            for col in range(1, 9):
                ws2.cell(row=row, column=col).fill = not_found_fill
            row += 1
    
    # 列幅調整
    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 25
    ws2.column_dimensions['D'].width = 18
    ws2.column_dimensions['E'].width = 22
    ws2.column_dimensions['F'].width = 18
    ws2.column_dimensions['G'].width = 50
    ws2.column_dimensions['H'].width = 25
    
    # ===== シート3: 俳優一覧 =====
    ws3 = wb.create_sheet("俳優一覧")
    
    headers3 = ["俳優名", "出演作品", "役名", "事務所", "SNS/備考"]
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 俳優データを集約
    actor_dict = {}
    for title, data in CAST_DATA.items():
        for actor in data["cast"]:
            name = actor["name"]
            if name not in actor_dict:
                actor_dict[name] = {
                    "titles": [],
                    "roles": [],
                    "agency": actor["agency"],
                    "social": actor["social"]
                }
            actor_dict[name]["titles"].append(title)
            actor_dict[name]["roles"].append(actor["role"])
    
    row = 2
    for name, info in sorted(actor_dict.items()):
        ws3.cell(row=row, column=1, value=name).border = thin_border
        ws3.cell(row=row, column=2, value=", ".join(info["titles"])).border = thin_border
        ws3.cell(row=row, column=3, value=", ".join(info["roles"])).border = thin_border
        ws3.cell(row=row, column=4, value=info["agency"]).border = thin_border
        ws3.cell(row=row, column=5, value=info["social"]).border = thin_border
        row += 1
    
    ws3.column_dimensions['A'].width = 20
    ws3.column_dimensions['B'].width = 40
    ws3.column_dimensions['C'].width = 30
    ws3.column_dimensions['D'].width = 20
    ws3.column_dimensions['E'].width = 60
    
    # ===== シート4: 事務所一覧 =====
    ws4 = wb.create_sheet("事務所一覧")
    
    headers4 = ["事務所名", "所属俳優", "出演作品数", "備考"]
    for col, header in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 事務所データを集約
    agency_dict = {}
    for name, info in actor_dict.items():
        agency = info["agency"]
        if agency and agency != "不明":
            if agency not in agency_dict:
                agency_dict[agency] = {"actors": [], "count": 0}
            agency_dict[agency]["actors"].append(name)
            agency_dict[agency]["count"] += len(info["titles"])
    
    row = 2
    agencies_info = [
        ("officeNAO", "みやなおこ", "http://officenao-info.com/"),
        ("eyes inc", "西野莉世", "若手俳優事務所、ショートドラマに多数出演"),
        ("Bianchetto", "新海航輝、秋葉七海", "https://bianchetto.co.jp/"),
        ("株式会社Quick", "紀乃はる", "https://919quick.com/"),
        ("nowhere film", "（自社制作）", "https://nowhere-film.jp/"),
    ]
    
    for agency, actors, note in agencies_info:
        ws4.cell(row=row, column=1, value=agency).border = thin_border
        ws4.cell(row=row, column=2, value=actors).border = thin_border
        ws4.cell(row=row, column=3, value=agency_dict.get(agency, {}).get("count", "-")).border = thin_border
        ws4.cell(row=row, column=4, value=note).border = thin_border
        row += 1
    
    ws4.column_dimensions['A'].width = 25
    ws4.column_dimensions['B'].width = 30
    ws4.column_dimensions['C'].width = 15
    ws4.column_dimensions['D'].width = 50
    
    # ===== シート5: 制作会社 =====
    ws5 = wb.create_sheet("制作会社情報")
    
    ws5['A1'] = "制作会社情報"
    ws5['A1'].font = Font(bold=True, size=14)
    
    ws5['A3'] = "nowhere film株式会社"
    ws5['A3'].font = Font(bold=True, size=12)
    ws5['A4'] = "所在地: 東京都渋谷区円山町28-8 第18宮廷マンション104"
    ws5['A5'] = "代表: 酒井大輝"
    ws5['A6'] = "設立: 2024年8月"
    ws5['A7'] = "事業: ショートドラマ制作専門プロダクション（500作品以上の制作実績）"
    ws5['A8'] = "HP: https://nowhere-film.jp/"
    ws5['A9'] = "Instagram: @nowhere_film_"
    ws5['A10'] = "TikTok: @nowhere_film"
    ws5['A11'] = "問い合わせ: contact@nowhere-film.jp"
    
    ws5['A13'] = "ReelShortとの関係"
    ws5['A13'].font = Font(bold=True, size=12)
    ws5['A14'] = "ReelShortの日本市場開拓におけるパートナーとして、グローバル人気作品の日本版リメイク等を担当"
    
    ws5.column_dimensions['A'].width = 80
    
    # 保存
    filename = f"ReelShort_日本オリジナル_キャスト情報_完全版_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = f"/home/kali/kali/reelshort/api_analysis/{filename}"
    wb.save(filepath)
    print(f"✅ Excelファイルを保存しました: {filename}")
    print(f"   パス: {filepath}")
    
    # サマリー表示
    print("\n" + "="*60)
    print("ReelShort日本オリジナル作品 キャスト情報 サマリー")
    print("="*60)
    print(f"総作品数: {total}")
    print(f"確認済み: {confirmed} 作品 ({confirmed/total*100:.1f}%)")
    print(f"確認済み俳優: {total_actors} 名")
    print("="*60)
    
    for title, data in CAST_DATA.items():
        if "✅" in data["status"]:
            print(f"\n✅ {title}")
            for actor in data["cast"]:
                print(f"   - {actor['name']} ({actor['role']}) @ {actor['agency']}")
    
    print("\n" + "="*60)
    print("未確認作品:")
    for title, data in CAST_DATA.items():
        if "❌" in data["status"]:
            print(f"   - {title}")
    
    return filename

if __name__ == "__main__":
    create_excel()

