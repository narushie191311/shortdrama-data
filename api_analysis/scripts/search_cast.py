#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ReelShort日本オリジナル作品のキャスト情報を収集するスクリプト
"""

import csv
import json
import time
import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 手動で収集したキャスト情報
MANUAL_CAST_DATA = {
    "冷酷御曹司の愛妻計画": {
        "actors": [
            {
                "name": "長田成哉",
                "role": "高嶺総一郎",
                "agency": "nowhere film所属俳優",
                "social": {"instagram": "", "twitter": ""},
                "source": "nowhere-film.jp"
            }
        ]
    },
    "財閥令嬢様の二重生活": {
        "actors": [
            {
                "name": "音野暁",
                "role": "主演",
                "agency": "不明",
                "social": {"instagram": "", "twitter": ""},
                "source": "nowhere-film.jp"
            },
            {
                "name": "江畑浩規",
                "role": "出演",
                "agency": "不明",
                "social": {"instagram": "", "twitter": ""},
                "source": "nowhere-film.jp"
            },
            {
                "name": "梁錦川",
                "role": "出演",
                "agency": "不明",
                "social": {"instagram": "", "twitter": ""},
                "source": "nowhere-film.jp"
            }
        ]
    },
    "一目惚れ！今すぐ結婚してくれますか？": {
        "actors": [
            {
                "name": "みやなおこ",
                "role": "出演",
                "agency": "officeNAO",
                "social": {
                    "instagram": "miyanaoko.com (INSTAGRAM)",
                    "twitter": "",
                    "website": "https://miyanaoko.com/"
                },
                "source": "miyanaoko.com"
            }
        ]
    }
}

# 作品リスト（CSVから読み込み）
TITLES = [
    "もう一度、君に恋をする",
    "消防士の元夫、悔恨の炎に焼かれて",
    "無職の夫は大富豪だった件!",
    "ダイヤモンドの再会",
    "銃と金とクリスマス",
    "財閥令嬢様の二重生活",
    "ケーキ職人のワタシが大富豪と偽装結婚した話",
    "禁断の誘い～彼のものに～",
    "Yesから始まるラブストーリー",
    "負け犬の私がバージンを捨てる！",
    "嫌いなアイツの専属メイド!?",
    "一目惚れ！今すぐ結婚してくれますか？",
    "冷酷御曹司の愛妻計画",
    "帰ってきたお嬢様",
    "せめて最後に愛のキスを",
    "奪われたプリマバレリーナ",
    "令嬢決戦！私こそが学園のクイーン"
]

def get_search_queries():
    """検索クエリを生成"""
    queries = {}
    for title in TITLES:
        queries[title] = [
            f"ReelShort {title} キャスト 俳優",
            f"ReelShort {title} 出演者",
            f"{title} 縦型ドラマ キャスト",
            f"nowhere film {title}",
            f"PR TIMES {title} ReelShort",
        ]
    return queries

def print_status():
    """収集状況を表示"""
    print("=" * 60)
    print("ReelShort日本オリジナル作品 キャスト情報収集状況")
    print("=" * 60)
    
    found_count = 0
    for title in TITLES:
        if title in MANUAL_CAST_DATA:
            actors = MANUAL_CAST_DATA[title]["actors"]
            print(f"✅ {title}")
            for actor in actors:
                print(f"   - {actor['name']} ({actor['role']}) @ {actor['agency']}")
            found_count += 1
        else:
            print(f"❌ {title} - 未調査")
    
    print("=" * 60)
    print(f"進捗: {found_count}/{len(TITLES)} 作品 ({found_count/len(TITLES)*100:.1f}%)")
    print("=" * 60)

def create_excel():
    """収集した情報をExcelに出力"""
    wb = Workbook()
    
    # スタイル定義
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    found_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    not_found_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # シート1: 作品別キャスト一覧
    ws1 = wb.active
    ws1.title = "作品別キャスト"
    
    headers = ["作品名", "調査状況", "俳優名", "役名", "事務所", "SNS/ウェブサイト", "情報ソース"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    row = 2
    for title in TITLES:
        if title in MANUAL_CAST_DATA:
            actors = MANUAL_CAST_DATA[title]["actors"]
            for i, actor in enumerate(actors):
                ws1.cell(row=row, column=1, value=title if i == 0 else "").border = thin_border
                ws1.cell(row=row, column=2, value="✅確認済み" if i == 0 else "").border = thin_border
                ws1.cell(row=row, column=3, value=actor["name"]).border = thin_border
                ws1.cell(row=row, column=4, value=actor["role"]).border = thin_border
                ws1.cell(row=row, column=5, value=actor["agency"]).border = thin_border
                
                # SNS情報
                social_info = []
                if actor.get("social"):
                    for platform, url in actor["social"].items():
                        if url:
                            social_info.append(f"{platform}: {url}")
                ws1.cell(row=row, column=6, value="\n".join(social_info) if social_info else "-").border = thin_border
                ws1.cell(row=row, column=7, value=actor.get("source", "-")).border = thin_border
                
                # 背景色
                for col in range(1, 8):
                    ws1.cell(row=row, column=col).fill = found_fill
                
                row += 1
        else:
            ws1.cell(row=row, column=1, value=title).border = thin_border
            ws1.cell(row=row, column=2, value="❌未調査").border = thin_border
            for col in range(3, 8):
                ws1.cell(row=row, column=col, value="-").border = thin_border
            for col in range(1, 8):
                ws1.cell(row=row, column=col).fill = not_found_fill
            row += 1
    
    # 列幅調整
    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 15
    ws1.column_dimensions['E'].width = 20
    ws1.column_dimensions['F'].width = 40
    ws1.column_dimensions['G'].width = 25
    
    # シート2: 俳優別一覧
    ws2 = wb.create_sheet("俳優別一覧")
    
    headers2 = ["俳優名", "事務所", "出演作品", "SNS/ウェブサイト"]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 俳優データを集約
    actor_dict = {}
    for title, data in MANUAL_CAST_DATA.items():
        for actor in data["actors"]:
            name = actor["name"]
            if name not in actor_dict:
                actor_dict[name] = {
                    "agency": actor["agency"],
                    "titles": [],
                    "social": actor.get("social", {})
                }
            actor_dict[name]["titles"].append(title)
    
    row = 2
    for name, info in actor_dict.items():
        ws2.cell(row=row, column=1, value=name).border = thin_border
        ws2.cell(row=row, column=2, value=info["agency"]).border = thin_border
        ws2.cell(row=row, column=3, value=", ".join(info["titles"])).border = thin_border
        
        social_info = []
        for platform, url in info["social"].items():
            if url:
                social_info.append(f"{platform}: {url}")
        ws2.cell(row=row, column=4, value="\n".join(social_info) if social_info else "-").border = thin_border
        row += 1
    
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 50
    ws2.column_dimensions['D'].width = 50
    
    # シート3: 検索用クエリ
    ws3 = wb.create_sheet("検索用クエリ")
    
    headers3 = ["作品名", "推奨検索クエリ1", "推奨検索クエリ2", "推奨検索クエリ3"]
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    row = 2
    queries = get_search_queries()
    for title, query_list in queries.items():
        ws3.cell(row=row, column=1, value=title).border = thin_border
        for i, query in enumerate(query_list[:3], 2):
            ws3.cell(row=row, column=i, value=query).border = thin_border
        row += 1
    
    ws3.column_dimensions['A'].width = 35
    ws3.column_dimensions['B'].width = 45
    ws3.column_dimensions['C'].width = 45
    ws3.column_dimensions['D'].width = 45
    
    # シート4: サマリー
    ws4 = wb.create_sheet("サマリー")
    
    ws4.cell(row=1, column=1, value="ReelShort 日本オリジナル作品 キャスト情報サマリー").font = Font(bold=True, size=16)
    ws4.cell(row=2, column=1, value=f"作成日時: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ws4.cell(row=4, column=1, value="統計情報").font = Font(bold=True, size=14)
    
    total_titles = len(TITLES)
    found_titles = len(MANUAL_CAST_DATA)
    total_actors = sum(len(data["actors"]) for data in MANUAL_CAST_DATA.values())
    
    ws4.cell(row=5, column=1, value=f"総作品数: {total_titles}")
    ws4.cell(row=6, column=1, value=f"調査済み作品数: {found_titles}")
    ws4.cell(row=7, column=1, value=f"未調査作品数: {total_titles - found_titles}")
    ws4.cell(row=8, column=1, value=f"確認済み俳優数: {total_actors}")
    ws4.cell(row=9, column=1, value=f"進捗率: {found_titles/total_titles*100:.1f}%")
    
    ws4.cell(row=11, column=1, value="調査リソース").font = Font(bold=True, size=14)
    resources = [
        "nowhere-film.jp - nowhere film公式サイト（多くの日本オリジナル作品を制作）",
        "prtimes.jp - PR TIMESでの配信開始プレスリリース",
        "各俳優の公式サイト/事務所サイト",
        "DuckDuckGo検索: \"ReelShort [作品名] キャスト\"",
        "reelshort.com/ja - ReelShort日本語版公式サイト",
    ]
    for i, resource in enumerate(resources, 12):
        ws4.cell(row=i, column=1, value=f"• {resource}")
    
    # 保存
    filename = f"ReelShort_日本オリジナル_キャスト情報_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(f"/home/kali/kali/reelshort/api_analysis/{filename}")
    print(f"\n📊 Excelファイルを保存しました: {filename}")
    return filename

if __name__ == "__main__":
    print_status()
    create_excel()

